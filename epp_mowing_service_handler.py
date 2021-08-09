"""
ALS/EPP = Friends For Ever
By Chris Hartley, Renew Indianapolis, Inc
(C)opyright 2019
GNU GPL License v3.0

Search Services in EPP to show only the desired service, export in Excel format.
Have invoice from ALS in ALS invoice format
Have directory of photos from ALS with files named in their particular way
Get a ZIP file with pictures and an Excel suitable for importing into EPP
for Attachments, Service Financials, and Service Property Detail

"""

from pathlib import Path, PurePath
from datetime import datetime, date
import zipfile
from openpyxl import Workbook, load_workbook
from os.path import expanduser
from math import ceil
import re
import fnmatch
import os


BUDGET_ACCUMULATOR_NAME = 'Maintenance.Renew'
ERRORS = []

def do_everything(status_box, progressbar, **kwargs):
    service_excel = kwargs['service_excel']
    invoice_excel = kwargs.get('invoice_excel', None)
    photo_path = kwargs['photo_path']
    output_zip_file = kwargs['output_zip_file']
    output_folder = kwargs['output_folder']
    budget_accumulator_name = kwargs['maintence_accumulator']

    p = Path(photo_path)

    attachments_header = [
        '',
        'Service Number',
        'External System Id',
        'Parcel Number',
        'Attachment Type',
        'Title',
        'Attachment Path',
    ]

    financials_header = [
        '',
        'Service Number',
        'External System Id',
        'Parcel Numbers',
        'Budget Accumulator Name',
        'Amount',
        'Amount Indicator',
        'Date Incurred',
        'Comment',
    ]

    property_details_header = [
        '',
        'Service Number',
        'External System Id',
        'Parcel Number',
        'Service Property.Comments',
        'Service Property.Completed By',
        'Service Property.Date Complete',
    ]


    attachments_header_text_fields = ['Parcel Number',]
    financials_header_text_fields = ['Parcel Numbers',]

    output_filename = 'ImportServiceTemplate-{0}.xlsx'.format(date.today(),)


    workbook = Workbook()
    workbook.remove_sheet(workbook.get_sheet_by_name('Sheet'))

    attachments_worksheet = workbook.create_sheet(title='Attachments')
    attachments_worksheet_row_counter = 1

    financials_worksheet = workbook.create_sheet(title='Service Financials')
    financials_worksheet_row_counter = 1

    property_detail_worksheet = workbook.create_sheet(title='Service Property Detail')
    property_detail_worksheet_row_counter = 1


    for i in range(1, len(attachments_header)):
        print('{} - attachments_worksheet'.format(i,))
        attachments_worksheet.cell(attachments_worksheet_row_counter, i, value=attachments_header[i])
    attachments_worksheet_row_counter = attachments_worksheet_row_counter + 1

    for i in range(1,len(financials_header)):
        financials_worksheet.cell(financials_worksheet_row_counter, i, value=financials_header[i])
    financials_worksheet_row_counter = financials_worksheet_row_counter + 1

    for i in range(1,len(property_details_header)):
        property_detail_worksheet.cell(property_detail_worksheet_row_counter, i, value=property_details_header[i])
    property_detail_worksheet_row_counter = property_detail_worksheet_row_counter + 1

    invoice_spreadsheet = load_workbook(filename=invoice_excel)
    invoice_tab = invoice_spreadsheet.active

    service_spreadsheet = load_workbook(filename = service_excel)
    service_tab = service_spreadsheet['Service Export']

    already_found = {}
    with zipfile.ZipFile(PurePath(output_folder, output_zip_file), 'w', zipfile.ZIP_DEFLATED) as myzip:
        status_box.setText('Starting...')
        num_records = max(enumerate(service_tab.rows))[0]
        step = int(ceil(num_records / 100))
        progress = 0 # % of progress bar filled
        for i,service_row in enumerate(service_tab.rows): # First we step through the service spreadsheet
            matched = False
            print(num_records, step, progress)
            if i == 0: # Skip the header row
                continue
            if i % step == 0:
                progress = progress + 1
                progressbar.setValue(progress)
            service_number = service_row[0].value
            parcel_number = service_row[3].value # this column can change depending on layout in epp service list view
            for i,row in enumerate(invoice_tab.rows): # look for matching parcels in the inovice spreadsheet
                invoice_parcel_number = row[0].value # subject to change as ALS changes invoice format
                invoice_address = row[1].value
                invoice_date_time = row[3].value
                invoice_amount = row[4].value

                if i == 0: # Skip header row
                    continue
                if str(invoice_parcel_number) == str(parcel_number):
                    matched = True
                    print(parcel_number)
                    if parcel_number in already_found:
                        already_found[parcel_number] = already_found[parcel_number] + 1
                        ERRORS.append('Parcel number {} included on invoice {} times'.format(parcel_number,already_found[parcel_number],))
                    else:
                        already_found[parcel_number] = 1
                    if type(invoice_date_time) == datetime:
                        date_incurred = invoice_date_time
                    elif invoice_date_time is not None:
                        date_incurred = datetime.strptime(invoice_date_time, '%m/%-d/%Y %-H:%M')
                    else:
                        if invoice_amount is not None and invoice_amount > 0:
                            ERRORS.append('Parcel number {} included on invoice for ${} but without a valid date/time listed.'.format(parcel_number,invoice_amount,))

                    # The problem is that the invoice links the parcel number to the street address
                    # that we use to find photos. If we don't have an invoice we need ePP street address to match mowing image file name address,
                    # which it often doesn't do exactly.
                #    images = list(p.glob('{}*.jpg'.format(row[1].value,)))
                    rule = re.compile(fnmatch.translate('{}*'.format(invoice_address,)), re.IGNORECASE)
                #    print('List Dir: {}'.format(os.listdir(photo_path),))
                    images = [name for name in os.listdir(photo_path) if rule.match(name)]
                    if len(images) == 0:
                        ERRORS.append('Parcel number {} included on invoice has no accompanying images'.format(parcel_number,))
                    for image in images:
                        print('Image Path - {}'.format(PurePath(photo_path, image),) )
                        attachments_worksheet.cell(attachments_worksheet_row_counter, 1, value=service_number) # Service Number
                        attachments_worksheet.cell(attachments_worksheet_row_counter, 2, value='') # External service number
                        attachments_worksheet.cell(attachments_worksheet_row_counter, 3, value=parcel_number) # Parcel Number
                        attachments_worksheet.cell(attachments_worksheet_row_counter, 4, value='Maintenance') # Service Type
                        attachments_worksheet.cell(attachments_worksheet_row_counter, 5, value=datetime.fromtimestamp(Path(PurePath(photo_path, image)).stat().st_mtime).strftime('%c')) # Title from image date
                        attachments_worksheet.cell(attachments_worksheet_row_counter, 6, value=str(PurePath(photo_path, image))) # Attachment Path
                        attachments_worksheet_row_counter = attachments_worksheet_row_counter + 1
                        myzip.write(Path(PurePath(photo_path, image)))

                    financials_worksheet.cell(financials_worksheet_row_counter, 1, value=service_number) # Service Number
                    financials_worksheet.cell(financials_worksheet_row_counter, 2, value='') # External System Id
                    financials_worksheet.cell(financials_worksheet_row_counter, 3, value=parcel_number) # Parcel Numbers
                    financials_worksheet.cell(financials_worksheet_row_counter, 4, value=budget_accumulator_name) # Budget Accumulator Name
                    financials_worksheet.cell(financials_worksheet_row_counter, 5, value=invoice_amount) # Amount
                    financials_worksheet.cell(financials_worksheet_row_counter, 6, value='C') # Amount Indicator - C=cost, I=income
                    financials_worksheet.cell(financials_worksheet_row_counter, 7, value=date_incurred) # Date Incurred
                    financials_worksheet.cell(financials_worksheet_row_counter, 8, value='Auto created') # Comment
                    financials_worksheet_row_counter = financials_worksheet_row_counter + 1

                    property_detail_worksheet.cell(property_detail_worksheet_row_counter, 1, value=service_number) # Service number
                    property_detail_worksheet.cell(property_detail_worksheet_row_counter, 2, value='') # External system id
                    property_detail_worksheet.cell(property_detail_worksheet_row_counter, 3, value=parcel_number) # Comment
                    property_detail_worksheet.cell(property_detail_worksheet_row_counter, 4, value='') # Comments
                    property_detail_worksheet.cell(property_detail_worksheet_row_counter, 5, value='') # Completed by
                    property_detail_worksheet.cell(property_detail_worksheet_row_counter, 6, value=date_incurred) # Date complete
                    property_detail_worksheet_row_counter = property_detail_worksheet_row_counter + 1

            if matched == False:
                print('matched == False, appending to error {}'.format(parcel_number,))
                ERRORS.append('Parcel number {} listed in service but not in invoice'.format(parcel_number,))

        invoice_parcel_numbers = [] # Generate list of all parcels in the invoice spreadsheet
        for i,row in enumerate(invoice_tab.rows):
            if i == 0: # Skip header row
                continue
            invoice_parcel_number = row[0].value # subject to change as ALS changes invoice format
            invoice_parcel_numbers.append(str(invoice_parcel_number))

        service_parcel_numbers = [] # Generate list of all parcels in the service spreadsheet
        for j,service_row in enumerate(service_tab.rows):
            if j == 0: # Skip the header row
                continue
            parcel_number = str(service_row[3].value)
            service_parcel_numbers.append(parcel_number)
        invoice_not_service_parcels = list(set(invoice_parcel_numbers) - set(service_parcel_numbers)) # calculate the difference
        for parcel in invoice_not_service_parcels:
            ERRORS.append('Parcel number {} listed in invoice but not in service'.format(parcel,))

        workbook.save(filename=PurePath(output_folder, output_filename) )
        myzip.write( PurePath(output_folder, output_filename), output_filename )
        progressbar.setValue(100)
        status_box.setText('Finished with {} error(s).'.format(len(ERRORS),))
        status_box.append('\n'.join(ERRORS))
        status_box.append('Saved {} and {}'.format(output_filename,output_zip_file,))
        print(ERRORS)

import sys
from PyQt5.QtWidgets import QApplication, QWidget, QInputDialog, QLineEdit, QFileDialog, QPushButton, QLabel, QProgressBar, QTextEdit
from PyQt5.QtWidgets import QComboBox
from PyQt5.QtGui import QIcon
from PyQt5.QtCore import pyqtSlot

class App(QWidget):
    service_excel = ''
    invoice_excel = ''
    photo_path = ''
    output_zip_file = 'service-update-{}.zip'.format(date.today(),)
    output_folder = expanduser("~")
    def __init__(self):
        super().__init__()
        self.title = 'ALS EPP FIXER - renewindianapolis.org'
        self.left = 10
        self.top = 10
        self.width = 640
        self.height = 600
        self.initUI()

    def initUI(self):
        self.setWindowTitle(self.title)
        self.setGeometry(self.left, self.top, self.width, self.height)

        self.welcome_text = QLabel('ALS -> EPP Mowing Import Creation Tool', self)
        self.welcome_text_sub = QLabel('By Chris Hartley, Renew Indianapolis', self)
        self.welcome_text.move(50,50)
        self.welcome_text_sub.move(50,70)
        service_button = QPushButton('Select Service Excel File', self)
        service_button.setToolTip('File exported from EPP')
        service_button.move(100,100)
        service_button.clicked.connect(self.on_click_service)
        self.service_button_lbl = QLabel('no file selected', self)
        self.service_button_lbl.setGeometry(100, 130, 500, 20)

        invoice_button = QPushButton('Select Invoice Excel File', self)
        invoice_button.setToolTip('Invoice File provided by ALS')
        invoice_button.move(100,160)
        invoice_button.clicked.connect(self.on_click_invoice)
        self.invoice_button_lbl = QLabel('no file selected', self)
        self.invoice_button_lbl.setGeometry(100, 190, 500, 20)

        picture_button = QPushButton('Select Directory with Pictures', self)
        picture_button.setToolTip('Folder with images provided by ALS')
        picture_button.move(100,220)
        picture_button.clicked.connect(self.on_click_folder)
        self.picture_button_lbl = QLabel('no folder selected', self)
        self.picture_button_lbl.setGeometry(100, 250, 500, 20)

        output_folder_button = QPushButton('Select Directory for output', self)
        output_folder_button.setToolTip('Folder to save ZIP and Excel files')
        output_folder_button.move(100, 280)
        output_folder_button.clicked.connect(self.on_click_output_folder)
        self.output_folder_button_lbl = QLabel(self.output_folder, self)
        self.output_folder_button_lbl.setGeometry(100, 310, 500, 20)

        maintence_accumulator_combo_box = QComboBox(self)
        maintence_accumulator_combo_box.addItem('Select value')
        maintence_accumulator_combo_box.addItem('Maintenance.Renew')
        maintence_accumulator_combo_box.addItem('Maintenance.DMD')

        maintence_accumulator_combo_box.move(100, 350)
        maintence_accumulator_combo_box.currentTextChanged.connect(self.maintenance_accumulator_onChanged)

        self.go_button = QPushButton('Go', self)
        self.go_button.move(400, 100)
        self.go_button.clicked.connect(self.on_click_go)


        self.exit_button = QPushButton('Exit', self)
        self.exit_button.move(400, 130)
        self.exit_button.clicked.connect(self.close)


        self.status_box = QTextEdit(self)
        self.status_box.move(100, 420)
        self.status_box.resize(400,100)
        self.status_box.setReadOnly(True)

        self.progress = QProgressBar(self)
        self.progress.setGeometry(100, 390, 410, 20)

        self.show()

    def openFileNameDialog(self, lbl):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        fileName, _ = QFileDialog.getOpenFileName(self,"Select Excel File", "", "All Files (*);;Excel Files (*.xlsx)" )
        if fileName:
            path_split = PurePath(fileName).parts
            lbl.setText(path_split[-1])
            return fileName

    def openFolderDialog(self, lbl):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        folder = QFileDialog.getExistingDirectory(self,"Select photo directory", "",QFileDialog.ShowDirsOnly)
        if folder:
            path_split = PurePath(folder).parts
            lbl.setText(path_split[-1])
            return folder

    @pyqtSlot()
    def on_click_service(self):
        self.service_excel = self.openFileNameDialog(lbl=self.service_button_lbl)
        #print(INPUT_SERVICE_EXCEL)

    @pyqtSlot()
    def on_click_invoice(self):
        self.invoice_excel = self.openFileNameDialog(lbl=self.invoice_button_lbl)

    @pyqtSlot()
    def on_click_folder(self):
        self.photo_path = self.openFolderDialog(lbl=self.picture_button_lbl)

    @pyqtSlot()
    def on_click_output_folder(self):
        self.output_folder = self.openFolderDialog(lbl=self.output_folder_button_lbl)

    #@pyqtSlot()
    def maintenance_accumulator_onChanged(self,text):
        print(text)
        self.maintenance_accumulator_value = text

    @pyqtSlot()
    def on_click_go(self):
        self.status_box.setText(self.invoice_excel)
        print('!!!!!',self.output_folder, self.output_zip_file, self.photo_path, self.invoice_excel, self.service_excel)
        do_everything(
            self.status_box, self.progress,
            output_folder=self.output_folder,
            output_zip_file=self.output_zip_file,
            photo_path=self.photo_path,
            invoice_excel=self.invoice_excel,
            service_excel=self.service_excel,
            maintence_accumulator=self.maintenance_accumulator_value,
            )



if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = App()
    sys.exit(app.exec_())
    #do_everything()
